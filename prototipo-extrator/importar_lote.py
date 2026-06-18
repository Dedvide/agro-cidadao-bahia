"""
Importador em lote — o funil de entrada de dados do SIPA.

Varre uma pasta e ingere os laudos:
  - .csv / .xlsx (tabela "longa")  → extração DETERMINÍSTICA (offline, sem chave)
  - .txt / .docx / .pdf            → extração com IA (precisa de ANTHROPIC_API_KEY)
  - .cdf / .mzml                   → adaptador de instrumento (stub, Fase 3)

Carrega no Data Lake e gera um RELATÓRIO (processados / aguardando IA / para revisão).
Validação: confiança < 0.7 → marcado para revisão humana antes de virar oficial.

Uso:  python importar_lote.py [pasta]      (padrão: exemplos/)
"""
from __future__ import annotations

import csv
import sys
import unicodedata
from pathlib import Path

from persistencia import conectar, salvar
from schema import Laudo, Resultado, TipoAmostra, TipoResultado, TipoSolicitante

AQUI = Path(__file__).parent
Q, I = TipoResultado.quantitativo, TipoResultado.identificacao

# aliases de cabeçalho (normalizados) → campo canônico
ALIAS = {
    "id": ["laudo", "amostra", "amostra_id", "protocolo", "laudo_id"],
    "municipio": ["municipio", "cidade"],
    "data": ["data_coleta", "data", "data_analise"],
    "categoria": ["categoria", "categoria_analise", "tipo_analise"],
    "analito": ["parametro", "analito", "ingrediente_ativo", "ensaio"],
    "valor": ["resultado", "valor"],
    "unidade": ["unidade", "und", "un"],
    "taxon": ["especie", "taxon", "praga"],
    "contagem": ["contagem", "quantidade", "qtd"],
    "estagio": ["estagio"],
    "propriedade": ["propriedade", "fazenda", "sitio"],
}
ANALITO_MAP = {"ph": "ph", "p": "fosforo", "k": "potassio", "ca": "calcio",
               "mg": "magnesio", "mo": "materia_organica"}
TIPO_AMOSTRA = {"monitoramento_pragas": TipoAmostra.inseto, "residuos_agrotoxicos": TipoAmostra.fruto,
                "fertilidade_solo": TipoAmostra.solo}


def _na(s) -> str:
    nf = unicodedata.normalize("NFKD", str(s).strip().lower())
    return "".join(c for c in nf if not unicodedata.combining(c))


def _float(v):
    """(valor, valor_texto) — trata vírgula decimal e '<0,01'."""
    if v is None or str(v).strip() == "":
        return None, None
    try:
        return float(str(v).replace(",", ".")), None
    except ValueError:
        return None, str(v).strip()


def _norm_analito(v):
    s = _na(v)
    return ANALITO_MAP.get(s, s.replace(" ", "_").replace("(", "").replace(")", ""))


def _ler_csv(caminho: Path) -> list[dict]:
    texto = caminho.read_text(encoding="utf-8")
    head = texto.splitlines()[0]
    sep = ";" if head.count(";") >= head.count(",") else ","
    return list(csv.DictReader(texto.splitlines(), delimiter=sep))


def _ler_xlsx(caminho: Path) -> list[dict]:
    from openpyxl import load_workbook
    ws = load_workbook(caminho, read_only=True, data_only=True).active
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return []
    header = [str(h).strip() if h is not None else "" for h in linhas[0]]
    return [dict(zip(header, r)) for r in linhas[1:]]


def _resolver_colunas(colunas: set[str]) -> dict:
    achados = {}
    for canon, aliases in ALIAS.items():
        for a in aliases:
            if a in colunas:
                achados[canon] = a
                break
    return achados


def parse_tabular(rows: list[dict]) -> list[Laudo]:
    """Converte uma tabela 'longa' (1 linha por resultado) em Laudos. [] se não reconhecida."""
    if not rows:
        return []
    rows = [{_na(k): v for k, v in r.items()} for r in rows]
    col = _resolver_colunas(set(rows[0].keys()))
    if "analito" not in col and "taxon" not in col:
        return []  # formato não reconhecido → roteia para IA

    grupos: dict[str, list[dict]] = {}
    for r in rows:
        chave = str(r.get(col.get("id", ""), "unico")) or "unico"
        grupos.setdefault(chave, []).append(r)

    laudos = []
    for chave, linhas in grupos.items():
        primeira = linhas[0]
        if "taxon" in col:
            categoria = "monitoramento_pragas"
        elif col.get("analito") == "ingrediente_ativo":
            categoria = "residuos_agrotoxicos"
        else:
            categoria = primeira.get(col.get("categoria", ""), "") or "fertilidade_solo"
        categoria = _na(categoria) if categoria else "fertilidade_solo"

        resultados = []
        for r in linhas:
            if "taxon" in col and r.get(col["taxon"]):
                cont, _ = _float(r.get(col.get("contagem", "")))
                resultados.append(Resultado(tipo=I, taxon=str(r[col["taxon"]]),
                                            contagem=int(cont) if cont is not None else None,
                                            estagio=r.get(col.get("estagio", ""))))
            elif "analito" in col and r.get(col["analito"]):
                val, txt = _float(r.get(col.get("valor", "")))
                resultados.append(Resultado(tipo=Q, analito=_norm_analito(r[col["analito"]]),
                                            valor=val, valor_texto=txt,
                                            unidade=r.get(col.get("unidade", "")) or None))
        if not resultados:
            continue
        laudos.append(Laudo(
            laudo_id=str(chave), categoria_analise=categoria,
            tipo_amostra=TIPO_AMOSTRA.get(categoria, TipoAmostra.outro),
            municipio=primeira.get(col.get("municipio", "")),
            propriedade=primeira.get(col.get("propriedade", "")),
            data_coleta=primeira.get(col.get("data", "")),
            resultados=resultados,
            laboratorio="CETAB", confianca_geral=0.9,
        ))
    return laudos


def _extrair_ia(caminho: Path) -> list[Laudo]:
    from extrator import carregar_texto, extrair
    return [extrair(carregar_texto(caminho))]


def importar_pasta(pasta: Path, conn) -> list[dict]:
    import os
    tem_chave = bool(os.environ.get("ANTHROPIC_API_KEY"))
    relatorio = []
    for arq in sorted(pasta.iterdir()):
        if not arq.is_file():
            continue
        ext = arq.suffix.lower()
        item = {"arquivo": arq.name, "tipo": ext.lstrip(".").upper() or "?", "status": "", "laudos": 0, "resultados": 0}
        try:
            if ext in (".csv", ".xlsx"):
                rows = _ler_csv(arq) if ext == ".csv" else _ler_xlsx(arq)
                laudos = parse_tabular(rows)
                if not laudos:  # formato não tabular → IA
                    if not tem_chave:
                        item["status"] = "formato livre → aguardando IA (defina ANTHROPIC_API_KEY)"
                        relatorio.append(item); continue
                    laudos = _extrair_ia(arq)
                    item["status"] = "OK (via IA)"
                else:
                    item["status"] = "OK (determinístico)"
            elif ext in (".txt", ".docx", ".pdf"):
                if not tem_chave:
                    item["status"] = "documento → aguardando IA (defina ANTHROPIC_API_KEY)"
                    relatorio.append(item); continue
                laudos = _extrair_ia(arq)
                item["status"] = "OK (via IA)"
            elif ext in (".cdf", ".mzml"):
                item["status"] = "arquivo de máquina → adaptador (Fase 3)"
                relatorio.append(item); continue
            else:
                item["status"] = "tipo não suportado"
                relatorio.append(item); continue

            for l in laudos:
                status = "rascunho" if l.confianca_geral < 0.7 else "importado"
                salvar(conn, l, status=status)
                item["laudos"] += 1
                item["resultados"] += len(l.resultados)
                if l.confianca_geral < 0.7:
                    item["status"] += " ⚠ revisão"
        except Exception as e:  # noqa: BLE001
            item["status"] = f"FALHA: {e}"
        relatorio.append(item)
    return relatorio


def main(argv) -> int:
    pasta = Path(argv[0]) if argv else (AQUI / "exemplos")
    if not pasta.is_dir():
        print(f"Pasta não encontrada: {pasta}")
        return 1
    saida = AQUI / "saida"; saida.mkdir(exist_ok=True)
    db = saida / "importacao_lote.sqlite"
    if db.exists():
        db.unlink()
    conn = conectar(db)

    rel = importar_pasta(pasta, conn)
    print(f"=== IMPORTAÇÃO EM LOTE · pasta: {pasta.name} ===")
    print(f"{'arquivo':32} {'tipo':6} status")
    print("-" * 78)
    for r in rel:
        extra = f" ({r['laudos']}L/{r['resultados']}R)" if r["laudos"] else ""
        print(f"{r['arquivo']:32} {r['tipo']:6} {r['status']}{extra}")
    print("-" * 78)
    ok = sum(1 for r in rel if r["laudos"])
    ia = sum(1 for r in rel if "aguardando IA" in r["status"])
    rev = sum(1 for r in rel if "revisão" in r["status"])
    print(f"Resumo: {ok} arquivo(s) importado(s) · {ia} aguardando IA · {rev} p/ revisão · "
          f"{conn.execute('SELECT COUNT(*) FROM laudo').fetchone()[0]} laudos no Data Lake")
    conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
